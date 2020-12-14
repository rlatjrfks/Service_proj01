import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np

def insert_excel_to_db():
    conn = pymysql.connect(
        host='ls-360d5e5827a35e0a46fa340307d68f5a00a3b151.cvbhe0hq8rxv.ap-northeast-2.rds.amazonaws.com', port=3306,
        user='dbmasteruser', passwd='Qa]HHh]dc1NsX>VLfo<=JA^1GcEWOCY$', db='dbmaster', charset='utf8')

    try:
        sql1 = 'select * from jongmok_list'
        cur1 = conn.cursor()
        cur1.execute(sql1)
        if sql1:
            with conn.cursor() as curs:

                sql = "update jongmok_list SET jongmok_listcol='%s', jongmok_listcol1='%s' where jongmok_name='%s'"

                wb = load_workbook("templates/jusikdata.xlsx", data_only=True)
                ws = wb['Sheet']

                iter_rows = iter(ws.rows)
                next(iter_rows)
                for row in iter_rows:
                    value = row[1].value.split(",")
                    vallist = ''.join(value)
                    value_float = row[2].value.split("%")
                    curs.execute(sql, (vallist, value_float[0], row[0].value))
                conn.commit()
        else:
            with conn.cursor() as curs:

                sql = 'insert into jongmok_list values(%s, %s, %s)'

                wb = load_workbook("templates/jusikdata.xlsx", data_only=True)
                ws = wb['Sheet']

                iter_rows = iter(ws.rows)
                next(iter_rows)
                for row in iter_rows:
                    value = row[1].value.split(",")
                    vallist = ''.join(value)
                    value_float = row[2].value.split("%")
                    curs.execute(sql, (row[0].value, vallist, value_float[0]))
                conn.commit()
    finally:
        conn.close()


if __name__ == "__main__":
    insert_excel_to_db()
